import customtkinter as ctk
from tkinter import messagebox, filedialog
import random
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.pagebreak import Break

# --- Configuration ---
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

# --- The Master List (85 Bilingual Items) ---
master_list = [
    "Adam / आदम", "Noah / नूह", "Abraham / अब्राहम", "Isaac / इसहाक", "Jacob / याकूब", 
    "Joseph / यूसुफ", "Moses / मूसा", "Aaron / हारून", "Joshua / यहोशू", "Samuel / शमूएल", 
    "David / दाऊद", "Solomon / सुलेमान", "Elijah / एलिय्याह", "Elisha / एलीशा", 
    "Isaiah / यशायाह", "Jeremiah / यिर्मयाह", "Daniel / दानिय्येल", "Peter / पतरस", 
    "Paul / पौलुस", "John / यूहन्ना", "Matthew / मत्ती",
    
    "Eve / हव्वा", "Sarah / सारा", "Rebekah / रिबका", "Rachel / राहेल", "Leah / लिआ", 
    "Miriam / मरियम", "Deborah / दबोरा", "Ruth / रूत", "Hannah / हन्ना", "Esther / एस्तेर", 
    "Mary (Mother) / मरियम (माता)", "Martha / मार्था", "Mary Magdalene / मरियम मगदलीनी", 
    "Elizabeth / इलीशिबा", "Priscilla / प्रिस्किला", "Dorcas / दौरकास", "Lydia / लुदिया", 
    "Naomi / नाओमी", "Rahab / राहाब", "Bathsheba / बतशेबा", "Jezebel / ईजेबेल",
    
    "Jerusalem / यरूशलेम", "Bethlehem / बेतलेहेम", "Nazareth / नासरत", "Jericho / यरीहो", 
    "Capernaum / कफरनहूम", "Bethany / बैतनिया", "Antioch / अन्ताकिया", "Damascus / दमिश्क", 
    "Ephesus / इफिसुस", "Corinth / कुरिन्थ", "Rome / रोम", "Babylon / बाबुल", 
    "Nineveh / नीनवे", "Eden / अदन", "Sodom / सदोम", "Canaan / कनान", "Egypt / मिस्र", 
    "Galilee / गलील", "Judea / यहूदिया", "Samaria / सामरिया", "Gethsemane / गतसमनी",
    
    "Mt. Sinai / सीनै पर्वत", "Mt. Zion / सिय्योन पर्वत", "Mt. Carmel / कर्मेल पर्वत", 
    "Mt. Nebo / नबो पर्वत", "Mt. Ararat / अरारात पर्वत", "Mt. Hermon / हेर्मोन पर्वत", 
    "Mt. Olives / जैतून पर्वत", "Mt. Tabor / ताबोर पर्वत", "Mt. Moriah / मोरिय्याह पर्वत", 
    "Jordan River / यरदन नदी", "Nile River / नील नदी", "Euphrates / फरात नदी", 
    "Tigris / हिद्देकेल नदी", "Sea of Galilee / गलील की झील", "Red Sea / लाल समुद्र", 
    "Dead Sea / मृत सागर", "Brook Kidron / किद्रोन नाला", "River Jabbok / यब्बोक नदी", 
    "Mt. Ebal / एबाल पर्वत", "Mt. Gerizim / गरिज्जीम पर्वत", "Mt. Gilboa / गिलबो पर्वत", 
    "River Chebar / कबार नदी"
]

class HousieApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Biblical Housie - 8 Tickets/Page")
        self.geometry("600x450")

        # --- Layout ---
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        # Header
        self.header = ctk.CTkFrame(self, fg_color="#2c3e50", corner_radius=0)
        self.header.grid(row=0, column=0, sticky="ew", ipady=15)
        
        ctk.CTkLabel(self.header, text="LANDSCAPE TICKET GENERATOR", 
                     font=("Roboto Medium", 20), text_color="white").pack()
        ctk.CTkLabel(self.header, text="8 Tickets per Page (2 Columns x 4 Rows)", 
                     text_color="#bdc3c7").pack()

        # Content
        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)
        
        self.info_lbl = ctk.CTkLabel(self.main_frame, 
                                     text="Ready to generate 100 tickets.\nFormat: Landscape Excel",
                                     font=("Arial", 14))
        self.info_lbl.pack(pady=30)

        # Button
        self.btn_gen = ctk.CTkButton(self.main_frame, text="GENERATE EXCEL (8/PAGE)", 
                                     command=self.generate_excel,
                                     width=250, height=50,
                                     font=("Arial", 16, "bold"),
                                     fg_color="#e67e22", hover_color="#d35400")
        self.btn_gen.pack(pady=10)

        # Footer
        ctk.CTkLabel(self, text="Developed by Yash Soloman & Shaan Jashwant", 
                     font=("Courier New", 12), text_color="#666").grid(row=2, column=0, pady=10)

    def generate_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="biblical_housie_landscape_8pp.xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Excel File"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "8 Tickets Per Page"
            
            # --- 1. PAGE SETUP (Landscape & Margins) ---
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False # Allow scrolling down/multiple pages
            
            # Narrow Margins to fit 8 tickets
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.25
            ws.page_margins.bottom = 0.25

            # --- 2. COLUMN WIDTHS ---
            # Grid: [Ticket1(A,B,C)] [Spacer(D)] [Ticket2(E,F,G)]
            col_width = 18 # Slightly narrower to fit 8 on landscape
            spacer_width = 15
            
            for col in ['A', 'B', 'C', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = col_width
            ws.column_dimensions['D'].width = spacer_width

            # --- 3. STYLES ---
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
            font_main = Font(name='Calibri', size=10) # Smaller font to fit items
            font_label = Font(name='Calibri', size=9, bold=True, color="555555")

            # --- 4. GENERATION LOOP ---
            random.seed()
            tickets = [random.sample(master_list, 9) for _ in range(100)]
            
            current_row = 1
            
            # We process 8 tickets at a time (4 rows of pairs)
            # i = 0, 8, 16, 24...
            for i in range(0, len(tickets), 8):
                
                # Inner loop for the 4 rows of tickets in this page
                for row_offset in range(0, 8, 2):
                    # Check if tickets exist
                    idx_left = i + row_offset
                    idx_right = i + row_offset + 1
                    
                    t_left = tickets[idx_left] if idx_left < len(tickets) else None
                    t_right = tickets[idx_right] if idx_right < len(tickets) else None

                    if t_left:
                        # --- LEFT TICKET ---
                        # Row 1
                        ws.cell(row=current_row, column=1, value=t_left[0])
                        ws.cell(row=current_row, column=2, value=t_left[1])
                        ws.cell(row=current_row, column=3, value=t_left[2])
                        # Row 2
                        ws.cell(row=current_row+1, column=1, value=t_left[3])
                        ws.cell(row=current_row+1, column=2, value=t_left[4])
                        ws.cell(row=current_row+1, column=3, value=t_left[5])
                        # Row 3
                        ws.cell(row=current_row+2, column=1, value=t_left[6])
                        ws.cell(row=current_row+2, column=2, value=t_left[7])
                        ws.cell(row=current_row+2, column=3, value=t_left[8])

                        # Styling Left
                        for r in range(current_row, current_row + 3):
                            for c in range(1, 4):
                                cell = ws.cell(row=r, column=c)
                                cell.border = thin_border
                                cell.alignment = centered
                                cell.font = font_main

                    if t_right:
                        # --- RIGHT TICKET ---
                        # Row 1
                        ws.cell(row=current_row, column=5, value=t_right[0])
                        ws.cell(row=current_row, column=6, value=t_right[1])
                        ws.cell(row=current_row, column=7, value=t_right[2])
                        # Row 2
                        ws.cell(row=current_row+1, column=5, value=t_right[3])
                        ws.cell(row=current_row+1, column=6, value=t_right[4])
                        ws.cell(row=current_row+1, column=7, value=t_right[5])
                        # Row 3
                        ws.cell(row=current_row+2, column=5, value=t_right[0]) # Copy error fixed logic below
                        ws.cell(row=current_row+2, column=5, value=t_right[6])
                        ws.cell(row=current_row+2, column=6, value=t_right[7])
                        ws.cell(row=current_row+2, column=7, value=t_right[8])

                        # Styling Right
                        for r in range(current_row, current_row + 3):
                            for c in range(5, 8):
                                cell = ws.cell(row=r, column=c)
                                cell.border = thin_border
                                cell.alignment = centered
                                cell.font = font_main

                    # --- MIDDLE LABEL ---
                    lbl_text = ""
                    if t_left: lbl_text += f"Ticket #{idx_left+1}"
                    if t_right: lbl_text += f"  |  #{idx_right+1}"
                    
                    lbl_cell = ws.cell(row=current_row+1, column=4, value=lbl_text)
                    lbl_cell.font = font_label
                    lbl_cell.alignment = centered

                    # Set Row Heights for this ticket pair
                    for r in range(current_row, current_row+3):
                        ws.row_dimensions[r].height = 35

                    # Move down for next pair + Spacer
                    current_row += 4 
                
                # --- PAGE BREAK LOGIC ---
                # After every 8 tickets (which is 4 pairs), we are at the end of a "page block".
                # current_row has incremented 4 times inside the loop.
                # We add a page break here.
                page_break_row = current_row - 1
                ws.row_breaks.append(Break(id=page_break_row))

            wb.save(file_path)
            messagebox.showinfo("Success", f"Excel file generated!\nSaved as: {os.path.basename(file_path)}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed:\n{str(e)}")

if __name__ == "__main__":
    app = HousieApp()
    app.mainloop()